"""
report_engine.py  —  OMRON Returns Reconciliation
===================================================
Cross-checks marketplace return/refund exports (Shopee, Lazada, TikTok) against
the OMRON Return Report tracker workbook, finds untracked return Order IDs, and
appends new rows enriched from a TC Order Report where available.

OMRON tracker main sheet:  "OMRON Return report - New Forma"
Header is on row 5 (0-indexed row 4). Data starts row 6.

Key columns (OMRON-specific):
  B  Marketplace/Platform
  C  Return Request Date
  D  Order Number
  E  Invoice Date
  F  Invoice Number
  G  RETURN LOGISTIC TRACKING NUMBER
  H  TRACKING STATUS          <- manual
  I  SKU
  J  Product Model (JDE Model)
  K  Qty
  L  GRASS Return Reason
  M  Return confirmation by MGL  <- manual
  N  Return Receiving date        <- manual
  O  If diff SKU received         <- manual
  P  Model Serial / Lot No.       <- manual
  Q  Fault Description            <- manual
  R  Action to be taken           <- manual
  S  Dispute to Marketplace       <- manual
  T  Redressing Instruction       <- manual
  U  Action Status After OMRON    <- manual
  V  STATUS                       <- manual
"""

from __future__ import annotations

import copy
import io
import re
import zipfile
import xml.etree.ElementTree as ET
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Optional

import openpyxl
import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------

@dataclass
class NewReturnRow:
    order_id: str
    platform: str
    return_request_date: Optional[datetime] = None
    invoice_date: Optional[datetime] = None
    invoice_number: Optional[str] = None
    tracking_number: Optional[str] = None
    sku: Optional[str] = None
    jde_model: Optional[str] = None
    qty: Any = None
    return_reason: Optional[str] = None
    source: str = ""
    notes: str = ""


@dataclass
class ReconciliationResult:
    marketplace: str
    total_in_file: int = 0
    already_tracked: list[str] = field(default_factory=list)
    new_rows: list[NewReturnRow] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def normalize_order_id(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else str(value)
    return str(value).strip()


def _coerce_datetime(value) -> Optional[datetime]:
    if value is None or value == "" or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, datetime):
        return value
    for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(str(value).strip(), fmt)
        except ValueError:
            pass
    try:
        return pd.to_datetime(value).to_pydatetime()
    except Exception:
        return None


def _safe_get(row: pd.Series, *names: str):
    for n in names:
        if n in row.index and pd.notna(row[n]):
            return row[n]
    return None


# ---------------------------------------------------------------------------
# File reading — magic-byte based, never relies on extension
# ---------------------------------------------------------------------------

def _read_raw(file_like) -> bytes:
    if isinstance(file_like, (bytes, bytearray)):
        return bytes(file_like)
    if hasattr(file_like, "read"):
        raw = file_like.read()
        if hasattr(file_like, "seek"):
            file_like.seek(0)
        return raw
    with open(file_like, "rb") as fh:
        return fh.read()


def _read_excel_any(file_like) -> pd.DataFrame:
    """
    Read .xls or .xlsx regardless of extension.
    Uses magic bytes: OLE2 (0xD0CF11E0) = .xls -> xlrd
                      ZIP  (PK\\x03\\x04)  = .xlsx -> openpyxl
    """
    raw = _read_raw(file_like)
    is_xls  = raw[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
    is_xlsx = raw[:4] == b"PK\x03\x04"

    if is_xls:
        return pd.read_excel(io.BytesIO(raw), engine="xlrd")
    if is_xlsx:
        try:
            return pd.read_excel(io.BytesIO(raw), engine="openpyxl")
        except Exception:
            return pd.read_excel(io.BytesIO(raw), engine="xlrd")
    try:
        return pd.read_excel(io.BytesIO(raw), engine="openpyxl")
    except Exception:
        return pd.read_excel(io.BytesIO(raw), engine="xlrd")


# ---------------------------------------------------------------------------
# Tracker loader
# ---------------------------------------------------------------------------

MANUAL_COLUMN_KEYWORDS = (
    "tracking \nstatus", "tracking status",
    "return confirmation",
    "return receiving date",
    "if diff sku",
    "model serial",
    "fault description",
    "action to be taken",
    "dispute",
    "redressing",
    "action status",
    "status",
)

FIELD_HEADER_CANDIDATES: dict[str, tuple[str, ...]] = {
    "platform":             ("marketplace/platform", "platform", "marketplace"),
    "return_request_date":  ("return request date",),
    "order_id":             ("order number", "order no", "order id"),
    "invoice_date":         ("invoice date", "ordered date", "order date"),
    "invoice_number":       ("invoice number", "invoice no"),
    "tracking_number":      ("return logistic tracking number", "tracking number", "logistic tracking"),
    "sku":                  ("sku",),
    "jde_model":            ("product model", "jde model"),
    "qty":                  ("qty", "quantity"),
    "return_reason":        ("grass return reason", "return reason", "reason"),
}


def _find_col(header_index: dict, *candidates: str) -> Optional[int]:
    for c in candidates:
        if c in header_index:
            return header_index[c]
    for key, i in header_index.items():
        for c in candidates:
            if c in key:
                return i
    return None


def load_tracker(file_like) -> dict:
    wb = openpyxl.load_workbook(file_like, data_only=False)

    # Find main sheet
    candidates = [s for s in wb.sheetnames if "return report" in s.lower()]
    sheet_name = candidates[0] if candidates else wb.sheetnames[0]
    ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))

    # Find header row — look for "Order Number"
    header_row_idx0 = 4  # default
    for i, row in enumerate(rows[:15]):
        cells = [str(c).strip().lower() for c in row if c is not None]
        if any("order number" in c or "order id" in c for c in cells):
            header_row_idx0 = i
            break

    header_row = rows[header_row_idx0]
    header_index = {str(h).strip().lower(): i for i, h in enumerate(header_row) if h is not None}

    order_col = _find_col(header_index, "order number", "order no", "order id")
    if order_col is None:
        raise ValueError(f"Cannot find Order Number column. Headers: {list(header_index.keys())}")

    existing_order_ids: set[str] = set()
    last_data_row_excel = header_row_idx0 + 1
    for offset, row in enumerate(rows[header_row_idx0 + 1:]):
        excel_row = header_row_idx0 + 2 + offset
        if any(v is not None for v in row):
            last_data_row_excel = excel_row
            val = row[order_col] if order_col < len(row) else None
            norm = normalize_order_id(val)
            if norm:
                existing_order_ids.add(norm)

    return {
        "workbook": wb,
        "sheet_name": sheet_name,
        "worksheet": ws,
        "header_row_idx0": header_row_idx0,
        "header_row_excel": header_row_idx0 + 1,
        "header": header_row,
        "header_index": header_index,
        "order_col": order_col,
        "existing_order_ids": existing_order_ids,
        "last_data_row_excel": last_data_row_excel,
    }


# ---------------------------------------------------------------------------
# Marketplace readers
# ---------------------------------------------------------------------------

def read_shopee_returns(file_like) -> pd.DataFrame:
    df = _read_excel_any(file_like)
    df.columns = [str(c).strip() for c in df.columns]
    if "Order ID" not in df.columns:
        raise ValueError(f"Shopee file missing 'Order ID'. Found: {list(df.columns)}")
    df["Order ID"] = df["Order ID"].apply(normalize_order_id)
    return df


def read_lazada_returns(file_like) -> pd.DataFrame:
    df = _read_excel_any(file_like)
    df.columns = [str(c).strip() for c in df.columns]
    if "Order ID" not in df.columns:
        raise ValueError(f"Lazada file missing 'Order ID'. Found: {list(df.columns)}")
    df["Order ID"] = df["Order ID"].apply(normalize_order_id)

    # Filter out Refunded rows — status col P = "Status" in header
    if "Status" in df.columns:
        df = df[df["Status"] != "Refunded"].reset_index(drop=True)

    return df


def read_tiktok_returns(file_like) -> pd.DataFrame:
    """
    TikTok exports have a structural defect where every cell sits in its own
    <row r='1'> element, causing openpyxl to see only the header row.
    We parse the ZIP/XML directly using cell refs (e.g. R5 = col R, row 5).
    Only 'In Process' and 'To Process' rows are returned.
    """
    KEEP_STATUSES = {"In Process", "To Process"}
    _NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"

    raw = _read_raw(file_like)

    # Try standard read first (non-malformed files)
    try:
        df_try = pd.read_excel(io.BytesIO(raw), engine="openpyxl")
        if len(df_try) > 0 and len(df_try.columns) > 1:
            df_try.columns = [str(c).strip() for c in df_try.columns]
            _standardise_tiktok_id(df_try)
            if "Return Status" in df_try.columns:
                df_try = df_try[df_try["Return Status"].isin(KEEP_STATUSES)]
            return df_try
    except Exception:
        pass

    # XML fallback
    with zipfile.ZipFile(io.BytesIO(raw)) as z:
        sheet_paths = [n for n in z.namelist() if "worksheets/" in n and n.endswith(".xml")]
        if not sheet_paths:
            raise ValueError("TikTok xlsx has no worksheet XML.")
        with z.open(sheet_paths[0]) as f:
            xml_content = f.read().decode("utf-8")

    root = ET.fromstring(xml_content)
    sd = root.find(f"{{{_NS}}}sheetData")
    if sd is None:
        return pd.DataFrame(columns=["Order ID"])

    cells: dict[int, dict[str, str]] = defaultdict(dict)
    for row_elem in sd.findall(f"{{{_NS}}}row"):
        for c_elem in row_elem.findall(f"{{{_NS}}}c"):
            ref = c_elem.get("r", "")
            m = re.match(r"([A-Z]+)(\d+)", ref)
            if m:
                col_letter, row_num = m.group(1), int(m.group(2))
                v = c_elem.find(f"{{{_NS}}}v")
                cells[row_num][col_letter] = v.text if v is not None else None

    if not cells:
        return pd.DataFrame(columns=["Order ID"])

    col_letters = sorted(
        {letter for row_dict in cells.values() for letter in row_dict},
        key=lambda x: (len(x), x),
    )
    col_names = [cells[1].get(l, l) for l in col_letters]

    records = []
    for row_num in sorted(k for k in cells if k > 1):
        records.append({col_names[i]: cells[row_num].get(l) for i, l in enumerate(col_letters)})

    df = pd.DataFrame(records)
    df.columns = [str(c).strip() for c in df.columns]

    status_col = next((c for c in df.columns if c.lower() == "return status"), None)
    if status_col:
        df = df[df[status_col].isin(KEEP_STATUSES)].reset_index(drop=True)

    _standardise_tiktok_id(df)
    return df


def _standardise_tiktok_id(df: pd.DataFrame) -> None:
    if "Order ID" in df.columns:
        df["Order ID"] = df["Order ID"].apply(normalize_order_id)
        return
    for c in df.columns:
        if "order id" in c.lower():
            df["Order ID"] = df[c].apply(normalize_order_id)
            return
    raise ValueError(f"TikTok file missing Order ID column. Found: {list(df.columns)}")


def read_tc_order_report(file_like) -> pd.DataFrame:
    raw = _read_raw(file_like)
    # CSV or Excel
    try:
        df = pd.read_csv(io.BytesIO(raw), dtype=str)
    except Exception:
        df = _read_excel_any(io.BytesIO(raw))
    df.columns = [c.strip() for c in df.columns]
    for c in df.columns:
        df[c] = df[c].astype(str).str.strip()
    if "order_id" not in df.columns:
        raise ValueError(f"TC Report missing 'order_id'. Found: {list(df.columns)}")
    df["order_id"] = df["order_id"].apply(normalize_order_id)
    return df


# ---------------------------------------------------------------------------
# Remove refunded/cancelled from tracker
# ---------------------------------------------------------------------------

SHOPEE_REMOVE_STATUSES = {"Request Refunded", "Request Cancelled", "Dispute Rejected", "Dispute Approved"}
SHOPEE_STATUS_PRIORITY = {
    "In Return": 0, "In Validation": 1, "In Seller Review": 2, "In Shopee Review": 3,
    "Dispute Approved": 4, "Request Refunded": 5, "Dispute Rejected": 6, "Request Cancelled": 7,
}


def remove_settled_rows(tracker: dict, shopee_df: Optional[pd.DataFrame], lazada_df: Optional[pd.DataFrame]) -> int:
    """Delete rows from the tracker that are now fully refunded/cancelled in the MP files."""
    ws: Worksheet = tracker["worksheet"]
    order_col = tracker["order_col"]
    rows = list(ws.iter_rows(values_only=True))

    # Build status maps
    sh_status: dict[str, str] = {}
    if shopee_df is not None and not shopee_df.empty and "Return / Refund Status" in shopee_df.columns:
        for oid, grp in shopee_df.groupby("Order ID"):
            statuses = grp["Return / Refund Status"].tolist()
            sh_status[oid] = sorted(statuses, key=lambda s: SHOPEE_STATUS_PRIORITY.get(s, 99))[0]

    lz_status: dict[str, list[str]] = {}
    if lazada_df is not None and not lazada_df.empty and "Status" in lazada_df.columns:
        for oid, grp in lazada_df.groupby("Order ID"):
            lz_status[oid] = grp["Status"].tolist()

    rows_to_delete = []
    for i, row in enumerate(rows[tracker["header_row_idx0"] + 1:], start=tracker["header_row_excel"] + 1):
        if any(v is not None for v in row):
            norm = normalize_order_id(row[order_col] if order_col < len(row) else None)
            if norm in sh_status and sh_status[norm] in SHOPEE_REMOVE_STATUSES:
                rows_to_delete.append(i)
            elif norm in lz_status and all(s == "Refunded" for s in lz_status[norm]):
                rows_to_delete.append(i)

    for row_num in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(row_num)

    return len(rows_to_delete)


# ---------------------------------------------------------------------------
# Reconcile one marketplace
# ---------------------------------------------------------------------------

def reconcile_marketplace(
    marketplace: str,
    mp_df: pd.DataFrame,
    existing_order_ids: set[str],
    tc_df: Optional[pd.DataFrame],
    shopee_status_map: Optional[dict] = None,
) -> ReconciliationResult:
    result = ReconciliationResult(marketplace=marketplace)

    if mp_df.empty or "Order ID" not in mp_df.columns:
        result.warnings.append(f"{marketplace} file has no usable data.")
        return result

    # Filter out settled statuses for Shopee
    if marketplace == "Shopee" and shopee_status_map:
        mp_df = mp_df[~mp_df["Order ID"].map(
            lambda x: shopee_status_map.get(x, "") in SHOPEE_REMOVE_STATUSES
        )]

    all_ids = mp_df["Order ID"].dropna().unique().tolist()
    result.total_in_file = len(all_ids)
    result.already_tracked = [oid for oid in all_ids if oid in existing_order_ids]
    new_ids = [oid for oid in all_ids if oid not in existing_order_ids]

    for order_id in new_ids:
        group = mp_df[mp_df["Order ID"] == order_id]
        first = group.iloc[0]

        tc_match = tc_df[tc_df["order_id"] == order_id] if tc_df is not None else pd.DataFrame()

        invoice, sku, notes = None, None, ""
        invoice_date_tc = None
        if not tc_match.empty:
            invoice = tc_match.iloc[0].get("invoice_number")
            # Use custom_sku (barcode format e.g. 4975479496318) NOT sku (EWG internal e.g. EWG0000034)
            custom_skus = [s for s in tc_match.get("custom_sku", pd.Series(dtype=str)).tolist() if s and s != "nan"]
            skus = custom_skus if custom_skus else []
            sku = " / ".join(dict.fromkeys(skus)) if skus else None
            # invoice_date from TC as fallback (TikTok has no order date in MP file)
            for tc_date_col in ("ordered_date", "order_date", "order_creation_date"):
                if tc_date_col in tc_match.columns:
                    v = tc_match.iloc[0].get(tc_date_col, "")
                    if v and str(v) not in ("", "nan"):
                        invoice_date_tc = _coerce_datetime(v)
                        break
        else:
            notes = "no TC match"

        if not sku:
            # Priority: Seller SKU ID (Lazada col J, barcode) → SKU (Shopee, barcode) → Seller SKU
            for col in ("Seller SKU ID", "SKU", "Seller SKU"):
                if col in group.columns:
                    vals = [str(s) for s in group[col].tolist() if pd.notna(s) and str(s) not in ("", "nan")]
                    if vals:
                        sku = " / ".join(dict.fromkeys(vals))
                        break

        # Invoice date: from MP file first, fall back to TC (especially for TikTok)
        invoice_date_mp = _coerce_datetime(_safe_get(first, "Order Creation Date", "Order Date"))
        invoice_date = invoice_date_mp or invoice_date_tc

        # Tracking: Lazada col R = "Tracking Number", Shopee = "Return Tracking Number"
        tracking = _safe_get(
            first,
            "Tracking Number",           # Lazada col R
            "Return Tracking Number",    # Shopee
            "Return Logistic Tracking Number",
            "Return Logistics Tracking ID",
        )

        result.new_rows.append(NewReturnRow(
            order_id=order_id,
            platform=f"{marketplace}-SG",
            return_request_date=_coerce_datetime(_safe_get(first, "Return Creation Time", "Return Order Date", "Time Requested")),
            invoice_date=invoice_date,
            invoice_number=invoice if invoice and invoice != "nan" else None,
            tracking_number=tracking,
            sku=sku,
            jde_model=None,
            qty=_safe_get(first, "Return Quantity", "Return Qty"),
            return_reason=_safe_get(first, "Return Reason"),
            source=marketplace,
            notes=notes,
        ))

    return result


# ---------------------------------------------------------------------------
# Append rows into tracker
# ---------------------------------------------------------------------------

def append_new_rows(tracker: dict, new_rows: list[NewReturnRow]) -> int:
    # Reload rows after potential deletions
    ws: Worksheet = tracker["worksheet"]
    rows2 = list(ws.iter_rows(values_only=True))
    last_row = tracker["header_row_excel"]
    for i, row in enumerate(rows2[tracker["header_row_idx0"] + 1:], start=tracker["header_row_excel"] + 1):
        if any(v is not None for v in row):
            last_row = i
    tracker["last_data_row_excel"] = last_row

    header_index = tracker["header_index"]
    ref_row = last_row
    n_cols = len(tracker["header"])

    field_col: dict[str, int] = {}
    for field_name, candidates in FIELD_HEADER_CANDIDATES.items():
        col0 = _find_col(header_index, *candidates)
        if col0 is not None:
            field_col[field_name] = col0 + 1

    start_row = last_row + 1
    for offset, row in enumerate(new_rows):
        excel_row = start_row + offset
        for col in range(1, n_cols + 1):
            src = ws.cell(row=ref_row, column=col)
            dst = ws.cell(row=excel_row, column=col)
            dst.font = copy.copy(src.font)
            dst.fill = copy.copy(src.fill)
            dst.border = copy.copy(src.border)
            dst.alignment = copy.copy(src.alignment)
            dst.number_format = src.number_format

        values = {
            "platform":            row.platform,
            "return_request_date": row.return_request_date,
            "order_id":            row.order_id,
            "invoice_date":        row.invoice_date,
            "invoice_number":      row.invoice_number,
            "tracking_number":     row.tracking_number,
            "sku":                 row.sku,
            "jde_model":           row.jde_model,
            "qty":                 row.qty,
            "return_reason":       row.return_reason,
        }
        for field_name, value in values.items():
            col = field_col.get(field_name)
            if col is not None:
                ws.cell(row=excel_row, column=col, value=value)

    return len(new_rows)


# ---------------------------------------------------------------------------
# Main orchestrator
# ---------------------------------------------------------------------------

def run_reconciliation(
    tracker_file,
    shopee_file=None,
    lazada_file=None,
    tiktok_file=None,
    tc_file=None,
) -> tuple[openpyxl.Workbook, list[ReconciliationResult]]:

    tracker = load_tracker(tracker_file)
    tc_df = read_tc_order_report(tc_file) if tc_file is not None else None

    shopee_df, lazada_df, tiktok_df = None, None, None
    shopee_status_map: dict[str, str] = {}

    if shopee_file is not None:
        shopee_df = read_shopee_returns(shopee_file)
        if "Return / Refund Status" in shopee_df.columns:
            for oid, grp in shopee_df.groupby("Order ID"):
                statuses = grp["Return / Refund Status"].tolist()
                shopee_status_map[oid] = sorted(statuses, key=lambda s: SHOPEE_STATUS_PRIORITY.get(s, 99))[0]

    if lazada_file is not None:
        lazada_df = read_lazada_returns(lazada_file)

    # Remove settled rows from tracker first
    removed = remove_settled_rows(tracker, shopee_df, lazada_df)

    # Reload existing IDs after deletions
    rows2 = list(tracker["worksheet"].iter_rows(values_only=True))
    order_col = tracker["order_col"]
    tracker["existing_order_ids"] = set()
    for row in rows2[tracker["header_row_idx0"] + 1:]:
        if any(v is not None for v in row):
            norm = normalize_order_id(row[order_col] if order_col < len(row) else None)
            if norm:
                tracker["existing_order_ids"].add(norm)

    results: list[ReconciliationResult] = []
    all_new_rows: list[NewReturnRow] = []

    if shopee_df is not None:
        res = reconcile_marketplace("Shopee", shopee_df, tracker["existing_order_ids"], tc_df, shopee_status_map)
        res.warnings.insert(0, f"{removed} refunded/cancelled rows removed from tracker.") if removed and not results else None
        results.append(res)
        all_new_rows.extend(res.new_rows)

    if lazada_df is not None:
        res = reconcile_marketplace("Lazada", lazada_df, tracker["existing_order_ids"], tc_df)
        results.append(res)
        all_new_rows.extend(res.new_rows)

    if tiktok_file is not None:
        tiktok_df = read_tiktok_returns(tiktok_file)
        res = reconcile_marketplace("TikTok", tiktok_df, tracker["existing_order_ids"], tc_df)
        results.append(res)
        all_new_rows.extend(res.new_rows)

    if removed and not results:
        results.append(ReconciliationResult(
            marketplace="—",
            warnings=[f"{removed} refunded/cancelled rows removed from tracker."]
        ))

    if all_new_rows:
        append_new_rows(tracker, all_new_rows)

    return tracker["workbook"], results, removed
